import os
import PyPDF2
import docx2txt
from Phone_Email_REGX.phone_email import *
import csv
import json
import openpyxl
import bs4

def word_file(filename):
    print("\n\ncontent from word file")
    doc = docx2txt.process(filename)
    extractor(doc, filename)

def text_file(filename):
    with open(filename, 'r') as fo:
        print("\n\ncontent from text file")
        extractor(fo.read(), filename)

def csv_file(filename):
    print("\n\ncontent from csv file")
    with open(filename, newline='') as f:
        reader = csv.reader(f, delimiter=':', quoting=csv.QUOTE_NONE)
        reader = csv.reader(f)
        data = ""
        for row in reader:
            data = data + " " + " ".join(row)
        extractor(data, filename)

def excel_file(filename):
    print("\n\ncontent from excel file")
    wb = openpyxl.load_workbook(filename)
    sheet_list = wb.sheetnames
    data = ""
    for i in sheet_list:
        wc = wb[i]
        tcol = wc.max_column + 1
        for i in range(1, wc.max_row + 1):
            for j in range(1, wc.max_column + 1):
                data = data + " " + str(wc.cell(row=i, column=j).value)
    extractor(str(data), filename)

def json_file(filename):
    f = open(filename)
    data_file = json.load(f)
    data = ""
    print("\n\ncontent from json file")
    for i in data_file['student_details']:
        data = data + " " + " ".join(list(i.values()))
    extractor(data, filename)

def html_file(filename):
    htmlfile = open(filename)
    soup = bs4.BeautifulSoup(htmlfile.read(), features="html.parser")
    elems = soup.select('p')
    print("\n\ncontent from html file")
    data = ""
    for i in range(len(elems)):
        data = data + " " + elems[i].getText()
    extractor(data, filename)

def pdf_file(filename):
    print("\n\ncontent from pdf file")
    pdfFileObj = open(filename, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    num_page = pdfReader.getNumPages()
    a = ""
    for i in range(0, num_page):
        pageObj = pdfReader.getPage(0)
        a = a + " " + pageObj.extractText()
    extractor(a, filename)


if __name__ == "__main__" :
    print("All the files in Data folder are")
    print(os.listdir('./Data'))
    os.chdir('./Data')
    for filename in os.listdir():
        if filename.endswith('.docx'):
            word_file(filename)
        elif filename.endswith('.txt'):
            text_file(filename)
        elif filename.endswith('.csv'):
            csv_file(filename)
        elif filename.endswith('.xlsx'):
            excel_file(filename)
        elif filename.endswith('.json'):
            json_file(filename)
        elif filename.endswith('.html'):
            html_file(filename)
        elif filename.endswith('.pdf'):
            pdf_file(filename)




